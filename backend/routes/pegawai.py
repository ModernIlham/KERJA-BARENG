from fastapi import APIRouter, HTTPException, Depends, Query
from typing import List, Optional, Dict, Any
from models import Pegawai, PegawaiCreate
from auth import get_current_user
from models import Pegawai, PegawaiCreate, MutasiPegawai, RiwayatKarir
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from datetime import datetime, timezone
import os
import math
from lib.activity_logger import log_activity

import pandas as pd
import uuid
from io import BytesIO
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File, Form
from lib.image_processor import process_image_upload
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

router = APIRouter()
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

@router.get("/pejabat")
async def get_pejabat_list(
    role: str = Query("PPK", description="Role to filter (e.g., PPK)"),
    current_user: str = Depends(get_current_user)
):
    # Filter pegawai where jabatan_melekat array contains the role (case insensitive)
    query = {"jabatan_melekat": {"$elemMatch": {"$regex": role, "$options": "i"}}}
    cursor = db.pegawai.find(query).sort("nama_lengkap", 1)
    items = await cursor.to_list(None)
    
    for item in items:
        if "_id" in item: item["_id"] = str(item["_id"])
        
    return items

@router.get("", response_model=Dict[str, Any])
async def get_pegawai_list(
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    status: Optional[str] = None,
    status_kepegawaian: Optional[str] = None,
    eselon1: Optional[str] = None,
    eselon2: Optional[str] = None,
    kategori_pegawai: Optional[str] = None,
    sort_by: Optional[str] = "nama_lengkap",
    sort_order: Optional[str] = "asc",
    current_user: str = Depends(get_current_user)
):
    skip = (page - 1) * limit
    query = {}
    
    # Enhanced search - search across multiple fields
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"nama_lengkap": search_regex},
            {"nip": search_regex},
            {"nik": search_regex},
            {"nrp": search_regex},
            {"email": search_regex},
            {"no_telp": search_regex},
            {"jabatan": search_regex},
            {"eselon1": search_regex},
            {"eselon2": search_regex},
            {"eselon3": search_regex},
            {"eselon4": search_regex},
            {"eselon5": search_regex},
            {"status_kepegawaian": search_regex},
            {"pangkat_golongan": search_regex},
            {"tempat_lahir": search_regex},
            {"nama_bank": search_regex},
            {"no_rekening": search_regex},
            {"nomor_kontrak": search_regex},
            {"nama_perusahaan": search_regex},
        ]
    
    # Filters
    if status:
        query["status"] = status
    if status_kepegawaian:
        query["status_kepegawaian"] = status_kepegawaian
    if eselon1:
        query["eselon1"] = eselon1
    if eselon2:
        query["eselon2"] = eselon2
    if kategori_pegawai:
        query["kategori_pegawai"] = kategori_pegawai
    
    # Sorting
    sort_direction = 1 if sort_order == "asc" else -1
    valid_sort_fields = ["nama_lengkap", "nip", "jabatan", "status", "status_kepegawaian", "eselon1", "pangkat_golongan", "tgl_lahir", "created_at"]
    if sort_by not in valid_sort_fields:
        sort_by = "nama_lengkap"
        
    total = await db.pegawai.count_documents(query)
    cursor = db.pegawai.find(query).skip(skip).limit(limit).sort(sort_by, sort_direction)
    items = await cursor.to_list(length=limit)
    
    # ObjectId to String
    for item in items:
        if "_id" in item: item["_id"] = str(item["_id"])
    
    return {
        "data": items,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": math.ceil(total / limit)
    }

@router.post("", response_model=Pegawai)
async def create_pegawai(pegawai_in: PegawaiCreate, current_user: dict = Depends(get_current_user)):
    # Check NIP conflict - ONLY if NIP is not empty
    if pegawai_in.nip and pegawai_in.nip.strip():
        existing = await db.pegawai.find_one({"nip": pegawai_in.nip})
        if existing:
            raise HTTPException(status_code=400, detail="NIP sudah digunakan oleh pegawai lain")
    
    # Check NIK conflict - ONLY if NIK is not empty
    if pegawai_in.nik and pegawai_in.nik.strip():
        existing_nik = await db.pegawai.find_one({"nik": pegawai_in.nik})
        if existing_nik:
            raise HTTPException(status_code=400, detail="NIK sudah digunakan oleh pegawai lain")
    
    # Check NRP conflict - ONLY if NRP is not empty
    if pegawai_in.nrp and pegawai_in.nrp.strip():
        existing_nrp = await db.pegawai.find_one({"nrp": pegawai_in.nrp})
        if existing_nrp:
            raise HTTPException(status_code=400, detail="NRP sudah digunakan oleh pegawai lain")
    
    # Handle pimpinan struktural auto-transfer before creating
    pegawai_data = pegawai_in.dict()
    if pegawai_data.get('is_pimpinan_struktural'):
        # Cari unit kerja tertinggi yang dipilih
        unit_key = None
        for es in ['eselon5', 'eselon4', 'eselon3', 'eselon2', 'eselon1']:
            if pegawai_data.get(es):
                unit_key = (es, pegawai_data.get(es))
                break
        
        if unit_key:
            # Reset pimpinan struktural lainnya di unit kerja yang sama
            await db.pegawai.update_many(
                {
                    unit_key[0]: unit_key[1],
                    "is_pimpinan_struktural": True
                },
                {"$set": {"is_pimpinan_struktural": False}}
            )
        
    new_pegawai = Pegawai(**pegawai_data)
    result = await db.pegawai.insert_one(new_pegawai.model_dump(by_alias=True, exclude=["id"]))
    
    # Log activity
    await log_activity(
        db=db,
        user_id=str(current_user.id),
        user_name=current_user.full_name or "Unknown",
        action="CREATE",
        module="Pegawai",
        target_id=str(result.inserted_id),
        details=f"Menambahkan pegawai baru: {pegawai_in.nama_lengkap}",
        metadata={"nama": pegawai_in.nama_lengkap, "nip": pegawai_in.nip}
    )
    
    return await db.pegawai.find_one({"_id": result.inserted_id})

@router.put("/{id}", response_model=Pegawai)
async def update_pegawai(id: str, pegawai_in: PegawaiCreate, current_user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    
    # Check NIP conflict if changed - ONLY if NIP is not empty
    if pegawai_in.nip and pegawai_in.nip.strip():
        existing = await db.pegawai.find_one({"nip": pegawai_in.nip, "_id": {"$ne": ObjectId(id)}})
        if existing: raise HTTPException(status_code=400, detail="NIP sudah digunakan oleh pegawai lain")
    
    # Check NIK conflict if changed - ONLY if NIK is not empty
    if pegawai_in.nik and pegawai_in.nik.strip():
        existing_nik = await db.pegawai.find_one({"nik": pegawai_in.nik, "_id": {"$ne": ObjectId(id)}})
        if existing_nik: raise HTTPException(status_code=400, detail="NIK sudah digunakan oleh pegawai lain")
    
    # Check NRP conflict if changed - ONLY if NRP is not empty
    if pegawai_in.nrp and pegawai_in.nrp.strip():
        existing_nrp = await db.pegawai.find_one({"nrp": pegawai_in.nrp, "_id": {"$ne": ObjectId(id)}})
        if existing_nrp: raise HTTPException(status_code=400, detail="NRP sudah digunakan oleh pegawai lain")
    
    update_data = pegawai_in.dict(exclude_unset=True)
    
    # Handle pimpinan struktural - hanya satu orang per unit kerja
    if update_data.get('is_pimpinan_struktural'):
        # Cari unit kerja tertinggi yang dipilih
        unit_key = None
        for es in ['eselon5', 'eselon4', 'eselon3', 'eselon2', 'eselon1']:
            if update_data.get(es):
                unit_key = (es, update_data.get(es))
                break
        
        if unit_key:
            # Reset pimpinan struktural lainnya di unit kerja yang sama
            await db.pegawai.update_many(
                {
                    "_id": {"$ne": ObjectId(id)},
                    unit_key[0]: unit_key[1],
                    "is_pimpinan_struktural": True
                },
                {"$set": {"is_pimpinan_struktural": False}}
            )
    
    res = await db.pegawai.find_one_and_update(
        {"_id": ObjectId(id)},
        {"$set": update_data},
        return_document=True
    )
    if not res: raise HTTPException(status_code=404)
    
    # Log activity
    await log_activity(
        db=db,
        user_id=str(current_user.id),
        user_name=current_user.full_name or "Unknown",
        action="UPDATE",
        module="Pegawai",
        target_id=id,
        details=f"Mengupdate data pegawai: {pegawai_in.nama_lengkap}"
    )
    
    return res

@router.get("/{id}", response_model=Pegawai)
async def get_pegawai_detail(id: str, current_user: str = Depends(get_current_user)):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    pegawai = await db.pegawai.find_one({"_id": ObjectId(id)})
    if not pegawai: raise HTTPException(status_code=404)
    return pegawai

@router.delete("/{id}")
async def delete_pegawai(id: str, current_user: dict = Depends(get_current_user)):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    
    # Get pegawai info before delete for logging
    pegawai = await db.pegawai.find_one({"_id": ObjectId(id)})
    
    res = await db.pegawai.delete_one({"_id": ObjectId(id)})
    if res.deleted_count == 0: raise HTTPException(status_code=404)
    
    # Log activity
    await log_activity(
        db=db,
        user_id=str(current_user.id),
        user_name=current_user.full_name or "Unknown",
        action="DELETE",
        module="Pegawai",
        target_id=id,
        details=f"Menghapus pegawai: {pegawai.get('nama_lengkap', 'Unknown') if pegawai else 'Unknown'}"
    )
    
    return {"message": "Pegawai deleted"}

@router.post("/{id}/mutasi", response_model=Pegawai)
async def mutasi_pegawai(id: str, mutasi: MutasiPegawai, current_user: str = Depends(get_current_user)):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    
    pegawai = await db.pegawai.find_one({"_id": ObjectId(id)})
    if not pegawai: raise HTTPException(status_code=404)
    
    # Create History Record
    riwayat = RiwayatKarir(
        jenis=mutasi.jenis_mutasi,
        deskripsi=mutasi.keterangan or f"Mutasi ke {mutasi.jabatan_baru}",
        jabatan_baru=mutasi.jabatan_baru,
        unit_kerja_baru=f"{mutasi.unit_kerja_baru.get('eselon1','')}, {mutasi.unit_kerja_baru.get('eselon2','')}",
        pangkat_baru=mutasi.pangkat_baru,
        sk_ref=mutasi.sk_ref,
        tanggal=mutasi.tgl_efektif
    )
    
    # Update Fields
    update_fields = {
        "jabatan": mutasi.jabatan_baru,
    }
    if mutasi.pangkat_baru:
        update_fields["pangkat_golongan"] = mutasi.pangkat_baru
        
    # Update Unit Kerja if provided
    if mutasi.unit_kerja_baru:
        for k, v in mutasi.unit_kerja_baru.items():
            if k in ['eselon1', 'eselon2', 'eselon3', 'eselon4']:
                update_fields[k] = v
                
    # Execute Update
    res = await db.pegawai.find_one_and_update(
        {"_id": ObjectId(id)},
        {
            "$set": update_fields,
            "$push": {"riwayat_karir": riwayat.dict()}
        },
        return_document=True
    )
    
@router.get("/import/template")
async def get_import_template(current_user: str = Depends(get_current_user)):
    """Generate Excel template with dropdown validations matching website options"""
    
    # --- REFERENCE DATA (Sesuai dengan PegawaiForm.js) ---
    PANGKAT_ASN = [
        "Juru Muda (I/a)", "Juru Muda Tingkat I (I/b)", "Juru (I/c)", "Juru Tingkat I (I/d)",
        "Pengatur Muda (II/a)", "Pengatur Muda Tingkat I (II/b)", "Pengatur (II/c)", "Pengatur Tingkat I (II/d)",
        "Penata Muda (III/a)", "Penata Muda Tingkat I (III/b)", "Penata (III/c)", "Penata Tingkat I (III/d)",
        "Pembina (IV/a)", "Pembina Tingkat I (IV/b)", "Pembina Utama Muda (IV/c)", "Pembina Utama Madya (IV/d)", "Pembina Utama (IV/e)"
    ]
    PANGKAT_PPPK = [
        "Golongan I", "Golongan II", "Golongan III", "Golongan IV", "Golongan V",
        "Golongan VI", "Golongan VII", "Golongan VIII", "Golongan IX", "Golongan X",
        "Golongan XI", "Golongan XII", "Golongan XIII", "Golongan XIV", "Golongan XV",
        "Golongan XVI", "Golongan XVII", "Golongan XVIII", "Golongan XIX"
    ]
    PANGKAT_TNI = [
        "Prajurit Dua", "Prajurit Satu", "Prajurit Kepala", 
        "Kopral Dua", "Kopral Satu", "Kopral Kepala",
        "Sersan Dua", "Sersan Satu", "Sersan Kepala", "Sersan Mayor",
        "Pembantu Letnan Dua", "Pembantu Letnan Satu",
        "Letnan Dua", "Letnan Satu", "Kapten",
        "Mayor", "Letnan Kolonel", "Kolonel",
        "Brigadir Jenderal", "Mayor Jenderal", "Letnan Jenderal", "Jenderal"
    ]
    PANGKAT_POLRI = [
        "Bhayangkara Dua", "Bhayangkara Satu", "Bhayangkara Kepala",
        "Ajun Brigadir Polisi Dua", "Ajun Brigadir Polisi Satu", "Ajun Brigadir Polisi",
        "Brigadir Polisi Dua", "Brigadir Polisi Satu", "Brigadir Polisi", "Brigadir Polisi Kepala",
        "Ajun Inspektur Polisi Dua", "Ajun Inspektur Polisi Satu",
        "Inspektur Polisi Dua", "Inspektur Polisi Satu", "Ajun Komisaris Polisi",
        "Komisaris Polisi", "Ajun Komisaris Besar Polisi", "Komisaris Besar Polisi"
    ]
    ALL_PANGKAT = PANGKAT_ASN + PANGKAT_PPPK + PANGKAT_TNI + PANGKAT_POLRI
    
    STATUS_KEPEGAWAIAN = ["PNS", "CPNS", "PPPK", "TNI", "POLRI", "Non-ASN", "Honorer"]
    JENIS_KELAMIN = ["Laki-laki", "Perempuan"]
    AGAMA = ["Islam", "Kristen", "Katolik", "Hindu", "Buddha", "Konghucu", "Lainnya"]
    STATUS_PERKAWINAN = ["Belum Kawin", "Kawin", "Cerai Hidup", "Cerai Mati"]
    PENDIDIKAN = ["SD", "SMP", "SMA/SMK", "D1", "D2", "D3", "D4/S1", "S2", "S3"]
    KEWARGANEGARAAN = ["WNI", "WNA"]
    STATUS_AKTIF = ["AKTIF", "CUTI", "TUGAS_BELAJAR", "KELUAR", "PENSIUN", "MUTASI KELUAR", "MENINGGAL"]
    # Non-ASN Fields
    JENIS_NON_ASN = ["Kontrak", "Outsourcing"]
    SUB_KATEGORI_NON_ASN = ["PPNPN", "Konsultan Individu", "Tenaga Ahli", "Teknisi", "Pramubakti", "Satpam", "Supir", "Magang"]
    
    # Status Fields
    STATUS_PENEMPATAN = ["Definitif", "Mutasi", "Penugasan"]
    STATUS_JABATAN = ["Definitif", "Plt", "Plh", "Pj", "Pjs"]
    JENIS_IDENTITAS_WNA = ["PASPOR", "KITAS", "KITAP"]
    
    # Kategori & Pimpinan - SESUAI UU ASN TERBARU
    KATEGORI_PEGAWAI = [
        "Jabatan Pimpinan Tinggi (JPT)",   # Eselon I & II
        "Jabatan Administrator",            # Eselon III
        "Jabatan Pengawas",                 # Eselon IV
        "Pejabat Pelaksana",                # Staf/Pelaksana
        "Jabatan Fungsional (JF)"           # Fungsional
    ]
    JENIS_PIMPINAN = ["Kepala", "Wakil"]
    YA_TIDAK = ["Ya", "Tidak"]
    
    # Fetch unit kerja from database
    units = await db.unit_kerja.find({}).to_list(1000)
    eselon1_list = sorted(list(set(u.get('nama_unit', '') for u in units if u.get('eselon') == '1' and u.get('nama_unit'))))
    eselon2_list = sorted(list(set(u.get('nama_unit', '') for u in units if u.get('eselon') == '2' and u.get('nama_unit'))))
    eselon3_list = sorted(list(set(u.get('nama_unit', '') for u in units if u.get('eselon') == '3' and u.get('nama_unit'))))
    eselon4_list = sorted(list(set(u.get('nama_unit', '') for u in units if u.get('eselon') == '4' and u.get('nama_unit'))))
    eselon5_list = sorted(list(set(u.get('nama_unit', '') for u in units if u.get('eselon') == '5' and u.get('nama_unit'))))
    
    # Fetch banks from database (dynamic)
    banks = await db.banks.find().sort("nama_bank", 1).to_list(1000)
    if not banks:
        # Use default if no banks in DB
        NAMA_BANK = ["BRI", "BNI", "Mandiri", "BTN", "Bank Syariah Indonesia (BSI)", "BCA", "CIMB Niaga", "Danamon", "Permata", "OCBC NISP", "Maybank", "Lainnya"]
    else:
        NAMA_BANK = [b.get('nama_bank') for b in banks if b.get('nama_bank')]
    
    # Create Workbook
    wb = Workbook()
    
    # --- SHEET 1: Template Import ---
    ws = wb.active
    ws.title = "Template Import"
    
    # Define ALL columns matching PegawaiForm.js - LENGKAP
    columns = [
        # A-H: Identitas Utama
        ("A", "Nama Lengkap", 30, True),
        ("B", "Gelar Depan", 12, False),
        ("C", "Gelar Belakang", 15, False),
        ("D", "Kewarganegaraan", 15, False),
        ("E", "NIP", 20, False),
        ("F", "NRP", 20, False),
        ("G", "NIK", 20, False),
        ("H", "NPWP", 22, False),
        
        # I-J: Identitas WNA
        ("I", "Jenis Identitas WNA", 18, False),
        ("J", "Nomor Identitas WNA", 20, False),
        
        # K-P: Data Pribadi
        ("K", "Jenis Kelamin", 15, False),
        ("L", "Tempat Lahir", 20, False),
        ("M", "Tanggal Lahir", 15, False),
        ("N", "Agama", 12, False),
        ("O", "Status Perkawinan", 18, False),
        ("P", "Pendidikan Terakhir", 20, False),
        
        # Q-V: Status Kepegawaian
        ("Q", "Status Kepegawaian", 18, False),
        ("R", "Pangkat/Golongan", 28, False),  # ASN + PPPK + TNI + POLRI
        ("S", "Status Penempatan", 18, False),
        ("T", "Instansi Asal", 25, False),
        ("U", "Masa Penugasan Berakhir", 22, False),
        ("V", "Status Jabatan", 15, False),
        
        # W-AA: Non-ASN Detail
        ("W", "Jenis Non-ASN", 15, False),
        ("X", "Sub-Kategori Non-ASN", 22, False),
        ("Y", "Nama Perusahaan (PT/CV)", 30, False),  # Untuk Outsourcing
        ("Z", "Nomor Kontrak", 25, False),  # Nomor kontrak aktif
        ("AA", "Tgl Mulai Kontrak", 18, False),
        ("AB", "Tgl Selesai Kontrak", 18, False),
        
        # AC-AN: Jabatan & Unit Kerja
        ("AC", "Jabatan Struktural", 35, False),
        ("AD", "Jabatan Fungsional Melekat", 30, False),
        ("AE", "Kategori Pegawai", 18, False),
        ("AF", "Pimpinan K/L", 15, False),  # Pimpinan Kementerian/Lembaga
        ("AG", "Jabatan Pimpinan K/L", 25, False),  # Menteri, Kepala Lembaga, dll
        ("AH", "Pimpinan Tertinggi", 18, False),
        ("AI", "Jenis Pimpinan", 15, False),
        ("AJ", "Eselon 1", 45, False),
        ("AK", "Eselon 2", 45, False),
        ("AL", "Eselon 3", 40, False),
        ("AM", "Eselon 4", 40, False),
        ("AN", "Eselon 5", 40, False),
        
        # AO-AR: Kontak & Bank
        ("AO", "No Telepon", 15, False),
        ("AP", "Email", 30, False),
        ("AQ", "Nama Bank", 28, False),
        ("AR", "No Rekening", 20, False),
        
        # AS-AT: Status & Lainnya
        ("AS", "Status Sistem", 15, False),
        ("AT", "Keterangan", 35, False),
    ]
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    required_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write Headers
    for col, (col_letter, header, width, is_required) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = required_fill if is_required else header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[col_letter].width = width
    
    # Add example row (row 2) - matching new LENGKAP structure
    example_data = [
        # A-H: Identitas Utama
        "Budi Santoso",  # A: Nama Lengkap
        "Dr.",  # B: Gelar Depan
        "S.E., M.M.",  # C: Gelar Belakang
        "WNI",  # D: Kewarganegaraan
        "198001012005011001",  # E: NIP
        "",  # F: NRP (untuk TNI/POLRI)
        "3201010101010001",  # G: NIK
        "12.345.678.9-012.000",  # H: NPWP
        
        # I-J: WNA (kosong untuk WNI)
        "",  # I: Jenis Identitas WNA
        "",  # J: Nomor Identitas WNA
        
        # K-P: Data Pribadi
        "Laki-laki",  # K: Jenis Kelamin
        "Jakarta",  # L: Tempat Lahir
        "01/01/1980",  # M: Tanggal Lahir
        "Islam",  # N: Agama
        "Kawin",  # O: Status Perkawinan
        "D4/S1",  # P: Pendidikan Terakhir
        
        # Q-V: Status Kepegawaian
        "PNS",  # Q: Status Kepegawaian
        "Penata (III/c)",  # R: Pangkat/Golongan (ASN/PPPK/TNI/POLRI)
        "Definitif",  # S: Status Penempatan
        "",  # T: Instansi Asal
        "",  # U: Masa Penugasan Berakhir
        "Definitif",  # V: Status Jabatan
        
        # W-AA: Non-ASN Detail (kosong untuk ASN)
        "",  # W: Jenis Non-ASN
        "",  # X: Sub-Kategori Non-ASN
        "",  # Y: Nama Perusahaan (PT/CV) - untuk Outsourcing
        "",  # Z: Nomor Kontrak
        "",  # AA: Tgl Mulai Kontrak
        "",  # AB: Tgl Selesai Kontrak
        
        # AC-AN: Jabatan & Unit Kerja
        "Kepala Seksi Umum",  # AC: Jabatan Struktural
        "PPK, Bendahara",  # AD: Jabatan Fungsional Melekat
        "Jabatan Pengawas",  # AE: Kategori Pegawai (sesuai UU ASN Terbaru)
        "Tidak",  # AF: Pimpinan K/L
        "",  # AG: Jabatan Pimpinan K/L (kosong jika bukan Pimpinan K/L)
        "Tidak",  # AH: Pimpinan Tertinggi
        "",  # AI: Jenis Pimpinan (kosong jika bukan pimpinan)
        eselon1_list[0] if eselon1_list else "",  # AJ: Eselon 1
        eselon2_list[0] if eselon2_list else "",  # AK: Eselon 2
        eselon3_list[0] if eselon3_list else "",  # AL: Eselon 3
        eselon4_list[0] if eselon4_list else "",  # AM: Eselon 4
        eselon5_list[0] if eselon5_list else "",  # AN: Eselon 5
        
        # AO-AR: Kontak & Bank
        "08123456789",  # AO: No Telepon
        "budi@example.com",  # AP: Email
        "BRI",  # AQ: Nama Bank
        "1234567890",  # AR: No Rekening
        
        # AS-AT: Status & Lainnya
        "AKTIF",  # AS: Status Sistem
        "",  # AT: Keterangan
    ]
    
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.fill = example_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Set row height
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    
    # --- ADD DATA VALIDATIONS (Dropdowns) ---
    def add_dropdown(col_letter, options, row_start=3, row_end=1000):
        if not options:
            return
        formula = '"' + ','.join(str(o) for o in options[:250]) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Pilih dari daftar yang tersedia"
        dv.errorTitle = "Input Tidak Valid"
        dv.prompt = "Pilih dari dropdown"
        dv.promptTitle = "Pilihan"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    
    # Add ALL validations matching PegawaiForm.js - LENGKAP
    add_dropdown("D", KEWARGANEGARAAN)  # Kewarganegaraan
    add_dropdown("I", JENIS_IDENTITAS_WNA)  # Jenis Identitas WNA
    add_dropdown("K", JENIS_KELAMIN)  # Jenis Kelamin
    add_dropdown("N", AGAMA)  # Agama
    add_dropdown("O", STATUS_PERKAWINAN)  # Status Perkawinan
    add_dropdown("P", PENDIDIKAN)  # Pendidikan Terakhir
    add_dropdown("Q", STATUS_KEPEGAWAIAN)  # Status Kepegawaian
    add_dropdown("R", ALL_PANGKAT)  # Pangkat/Golongan (ASN + PPPK + TNI + POLRI)
    add_dropdown("S", STATUS_PENEMPATAN)  # Status Penempatan
    add_dropdown("V", STATUS_JABATAN)  # Status Jabatan
    add_dropdown("W", JENIS_NON_ASN)  # Jenis Non-ASN
    add_dropdown("X", SUB_KATEGORI_NON_ASN)  # Sub-Kategori Non-ASN
    add_dropdown("AE", KATEGORI_PEGAWAI)  # Kategori Pegawai
    add_dropdown("AF", YA_TIDAK)  # Pimpinan K/L
    add_dropdown("AG", ["Menteri", "Wakil Menteri", "Kepala Lembaga", "Wakil Kepala Lembaga", "Kepala Badan", "Wakil Kepala Badan", "Direktur Utama", "Komisaris Utama"])  # Jabatan Pimpinan K/L
    add_dropdown("AH", YA_TIDAK)  # Pimpinan Tertinggi
    add_dropdown("AI", JENIS_PIMPINAN)  # Jenis Pimpinan
    add_dropdown("AQ", NAMA_BANK)  # Nama Bank
    add_dropdown("AS", STATUS_AKTIF)  # Status Sistem
    
    # Unit kerja dropdowns
    if eselon1_list:
        add_dropdown("AJ", eselon1_list)
    if eselon2_list:
        add_dropdown("AK", eselon2_list)
    if eselon3_list:
        add_dropdown("AL", eselon3_list)
    if eselon4_list:
        add_dropdown("AM", eselon4_list)
    if eselon5_list:
        add_dropdown("AN", eselon5_list)
    
    # --- SHEET 2: Referensi Data ---
    ws_ref = wb.create_sheet("Referensi Data")
    ws_ref.sheet_properties.tabColor = "0070C0"
    
    ref_data = [
        ("Status Kepegawaian", STATUS_KEPEGAWAIAN),
        ("Pangkat ASN (PNS/CPNS)", PANGKAT_ASN),
        ("Golongan PPPK (I-XIX)", PANGKAT_PPPK),
        ("Pangkat TNI", PANGKAT_TNI),
        ("Pangkat POLRI", PANGKAT_POLRI),
        ("Jenis Non-ASN", JENIS_NON_ASN),
        ("Sub-Kategori Non-ASN", SUB_KATEGORI_NON_ASN),
        ("Status Penempatan", STATUS_PENEMPATAN),
        ("Status Jabatan", STATUS_JABATAN),
        ("Kategori Pegawai", KATEGORI_PEGAWAI),
        ("Jenis Pimpinan", JENIS_PIMPINAN),
        ("Ya/Tidak", YA_TIDAK),
        ("Jenis Kelamin", JENIS_KELAMIN),
        ("Agama", AGAMA),
        ("Status Perkawinan", STATUS_PERKAWINAN),
        ("Pendidikan", PENDIDIKAN),
        ("Kewarganegaraan", KEWARGANEGARAAN),
        ("Jenis ID WNA", JENIS_IDENTITAS_WNA),
        ("Nama Bank", NAMA_BANK),
        ("Status Sistem", STATUS_AKTIF),
        ("Eselon 1", eselon1_list),
        ("Eselon 2", eselon2_list),
        ("Eselon 3", eselon3_list),
        ("Eselon 4", eselon4_list),
        ("Eselon 5", eselon5_list),
    ]
    
    col_offset = 1
    for title, items in ref_data:
        # Header
        cell = ws_ref.cell(row=1, column=col_offset, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
        # Items
        for row_idx, item in enumerate(items, 2):
            ws_ref.cell(row=row_idx, column=col_offset, value=item)
        
        # Set width dynamically
        max_len = max(len(str(title)), max((len(str(i)) for i in items), default=10))
        ws_ref.column_dimensions[get_column_letter(col_offset)].width = min(max_len + 5, 50)
        col_offset += 1
    
    # --- SHEET 3: Petunjuk ---
    ws_help = wb.create_sheet("Petunjuk Pengisian")
    ws_help.sheet_properties.tabColor = "00B050"
    
    instructions = [
        ("PETUNJUK IMPORT DATA PEGAWAI", ""),
        ("", ""),
        ("KOLOM WAJIB (KUNING)", ""),
        ("- Nama Lengkap", "Wajib diisi untuk semua pegawai"),
        ("", ""),
        ("FORMAT TANGGAL", ""),
        ("- Gunakan format DD/MM/YYYY", "Contoh: 01/01/1980"),
        ("", ""),
        ("DROPDOWN", ""),
        ("- Kolom dengan dropdown akan menampilkan pilihan saat diklik", ""),
        ("- Lihat sheet 'Referensi Data' untuk daftar lengkap pilihan", ""),
        ("", ""),
        ("PANDUAN PER STATUS KEPEGAWAIAN:", ""),
        ("", ""),
        ("PNS / CPNS:", ""),
        ("- Isi kolom NIP", "198001012005011001"),
        ("- Isi kolom Pangkat/Golongan", "Juru Muda (I/a) s.d. Pembina Utama (IV/e)"),
        ("", ""),
        ("PPPK:", ""),
        ("- Isi kolom NIP", "Sama format dengan PNS"),
        ("- Isi kolom Pangkat/Golongan", "Golongan I s.d. Golongan XIX"),
        ("", ""),
        ("TNI / POLRI:", ""),
        ("- Isi kolom NRP", ""),
        ("- Isi kolom Pangkat/Golongan sesuai jenjang", ""),
        ("", ""),
        ("Non-ASN / Honorer:", ""),
        ("- Isi kolom NIK", ""),
        ("- Isi kolom Jenis Non-ASN & Sub-Kategori", ""),
        ("- Jika Outsourcing, isi kolom Nama Perusahaan (PT/CV)", ""),
        ("- Isi tanggal mulai & selesai kontrak", ""),
        ("", ""),
        ("WNA (Warga Negara Asing):", ""),
        ("- Pilih Kewarganegaraan = WNA", ""),
        ("- Isi Jenis Identitas WNA (PASPOR/KITAS/KITAP)", ""),
        ("- Isi Nomor Identitas WNA", ""),
        ("", ""),
        ("UNIT KERJA:", ""),
        ("- Pilih dari Eselon 1 sampai Eselon 5 berurutan", ""),
        ("- Data dropdown diambil dari database", ""),
        ("", ""),
        ("PIMPINAN KEMENTERIAN/LEMBAGA:", ""),
        ("- Pimpinan K/L: Ya/Tidak (di atas struktur Eselon I)", ""),
        ("- Jika Ya, pilih jabatan: Menteri/Wakil Menteri/Kepala Lembaga/dll", ""),
        ("", ""),
        ("KATEGORI & PIMPINAN (SESUAI UU ASN TERBARU):", ""),
        ("- Jabatan Pimpinan Tinggi (JPT) = Eselon I & II", ""),
        ("- Jabatan Administrator = Eselon III", ""),
        ("- Jabatan Pengawas = Eselon IV", ""),
        ("- Pejabat Pelaksana = Staf/Pelaksana", ""),
        ("- Jabatan Fungsional (JF) = Fungsional", ""),
        ("- Pimpinan Tertinggi: Ya/Tidak", ""),
        ("- Jika Ya, pilih Jenis Pimpinan: Kepala/Wakil", ""),
        ("", ""),
        ("TIPS PENTING:", ""),
        ("- Hapus baris contoh (baris 2 hijau) sebelum import", ""),
        ("- Simpan file dalam format .xlsx", ""),
        ("- Data dengan NIP/NIK/NRP duplikat akan dilewati", ""),
    ]
    
    for row_idx, (col1, col2) in enumerate(instructions, 1):
        cell1 = ws_help.cell(row=row_idx, column=1, value=col1)
        cell2 = ws_help.cell(row=row_idx, column=2, value=col2)
        if "PETUNJUK" in col1:
            cell1.font = Font(bold=True, size=14, color="1F4E79")
        elif col1.endswith(":") and not col1.startswith("-"):
            cell1.font = Font(bold=True, color="1F4E79")
        elif col1.startswith("-"):
            cell1.font = Font(italic=False)
    
    ws_help.column_dimensions['A'].width = 45
    ws_help.column_dimensions['B'].width = 55
    
    # Freeze panes
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_pegawai.xlsx"}
    )

@router.post("/import")
async def import_pegawai(
    file: UploadFile = File(...), 
    current_user: str = Depends(get_current_user)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Format file harus Excel (.xlsx, .xls)")
        
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), sheet_name=0)  # Read first sheet only
        
        # 1. Required columns (minimal)
        required_columns = {"NIP", "Nama Lengkap"}
        file_cols = [c.strip() for c in df.columns]
        missing_required = [c for c in required_columns if c not in file_cols]
        
        if missing_required:
            raise HTTPException(status_code=400, detail=f"Kolom wajib tidak ditemukan: {', '.join(missing_required)}")
            
        # 2. Process Data
        success_count = 0
        skipped_count = 0
        failed_count = 0
        errors = []
        
        # Clean data: Trim whitespace from all string columns
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        
        # Helper to get cell value
        def get_val(row, col, default=None):
            val = row.get(col)
            if pd.isna(val) or str(val).lower() in ['nan', 'none', '']:
                return default
            return str(val).strip() if isinstance(val, (str, int, float)) else default
        
        # Iterate rows
        for index, row in df.iterrows():
            try:
                # Basic Validation - Nama Lengkap is required
                nama = get_val(row, "Nama Lengkap")
                if not nama or "Budi Santoso" in nama:  # Skip example row
                    if not nama:
                        failed_count += 1
                        errors.append(f"Baris {index+2}: Nama Lengkap kosong")
                    continue

                # Get identity fields
                nip = get_val(row, "NIP")
                nrp = get_val(row, "NRP")
                nik = get_val(row, "NIK")
                npwp = get_val(row, "NPWP")
                
                # Skip example rows
                if nip and "198001012005011001" in nip:
                    continue

                # Uniqueness Check
                query_conditions = []
                if nip: query_conditions.append({"nip": nip})
                if nrp: query_conditions.append({"nrp": nrp})
                if nik: query_conditions.append({"nik": nik})
                if npwp: query_conditions.append({"npwp": npwp})
                
                if query_conditions:
                    existing = await db.pegawai.find_one({"$or": query_conditions})
                    if existing:
                        skipped_count += 1
                        continue
                
                # Parse date fields helper
                def parse_date(date_str):
                    if not date_str:
                        return None
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"]:
                        try:
                            return datetime.strptime(date_str, fmt).isoformat()
                        except:
                            continue
                    return None
                    
                # Construct Object with ALL fields matching PegawaiForm.js
                new_pegawai = Pegawai(
                    # Identitas Utama
                    nama_lengkap=nama,
                    gelar_depan=get_val(row, "Gelar Depan"),
                    gelar_belakang=get_val(row, "Gelar Belakang"),
                    kewarganegaraan=get_val(row, "Kewarganegaraan", "WNI"),
                    nip=nip,
                    nrp=nrp,
                    nik=nik,
                    npwp=npwp,
                    
                    # WNA Identity
                    jenis_identitas_wna=get_val(row, "Jenis Identitas WNA"),
                    nomor_identitas_wna=get_val(row, "Nomor Identitas WNA"),
                    
                    # Data Pribadi
                    jenis_kelamin=get_val(row, "Jenis Kelamin"),
                    tempat_lahir=get_val(row, "Tempat Lahir"),
                    tanggal_lahir=parse_date(get_val(row, "Tanggal Lahir")),
                    agama=get_val(row, "Agama"),
                    status_perkawinan=get_val(row, "Status Perkawinan"),
                    pendidikan_terakhir=get_val(row, "Pendidikan Terakhir"),
                    
                    # Status Kepegawaian
                    status_kepegawaian=get_val(row, "Status Kepegawaian"),
                    pangkat_golongan=get_val(row, "Pangkat/Golongan") or get_val(row, "Pangkat/Golongan ASN") or get_val(row, "Golongan PPPK"),
                    status_penempatan=get_val(row, "Status Penempatan"),
                    instansi_asal=get_val(row, "Instansi Asal"),
                    masa_penugasan_end=parse_date(get_val(row, "Masa Penugasan Berakhir")),
                    status_jabatan=get_val(row, "Status Jabatan"),
                    
                    # Non-ASN Detail
                    jenis_non_asn=get_val(row, "Jenis Non-ASN"),
                    sub_kategori_non_asn=get_val(row, "Sub-Kategori Non-ASN"),
                    nama_perusahaan=get_val(row, "Nama Perusahaan (PT/CV)"),
                    nomor_kontrak=get_val(row, "Nomor Kontrak"),
                    tgl_mulai_kontrak=parse_date(get_val(row, "Tgl Mulai Kontrak")),
                    tgl_selesai_kontrak=parse_date(get_val(row, "Tgl Selesai Kontrak")),
                    
                    # Jabatan & Unit Kerja
                    jabatan=get_val(row, "Jabatan Struktural") or get_val(row, "Jabatan"),
                    jabatan_melekat=get_val(row, "Jabatan Fungsional Melekat"),
                    kategori_pegawai=get_val(row, "Kategori Pegawai"),
                    is_pimpinan_kl=(get_val(row, "Pimpinan K/L") == "Ya"),
                    jabatan_pimpinan_kl=get_val(row, "Jabatan Pimpinan K/L"),
                    is_pimpinan_tertinggi=(get_val(row, "Pimpinan Tertinggi") == "Ya"),
                    jenis_pimpinan=get_val(row, "Jenis Pimpinan"),
                    eselon1=get_val(row, "Eselon 1"),
                    eselon2=get_val(row, "Eselon 2"),
                    eselon3=get_val(row, "Eselon 3"),
                    eselon4=get_val(row, "Eselon 4"),
                    eselon5=get_val(row, "Eselon 5"),
                    
                    # Kontak & Bank
                    no_telp=get_val(row, "No Telepon") or get_val(row, "No Telp"),
                    email=get_val(row, "Email"),
                    nama_bank=get_val(row, "Nama Bank"),
                    no_rekening=get_val(row, "No Rekening"),
                    
                    # Status & Lainnya
                    status=get_val(row, "Status Sistem") or get_val(row, "Status", "AKTIF"),
                    keterangan=get_val(row, "Keterangan")
                )
                
                # Insert
                await db.pegawai.insert_one(new_pegawai.model_dump(by_alias=True, exclude=["id"]))
                success_count += 1
                
            except Exception as row_err:
                failed_count += 1
                errors.append(f"Baris {index+2}: {str(row_err)}")
                
        return {
            "message": "Import selesai",
            "success": success_count,
            "skipped": skipped_count,
            "failed": failed_count,
            "errors": errors[:50]  # Limit error messages
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@router.post("/{id}/signature")
async def upload_signature(
    id: str,
    file: UploadFile = File(...),
    current_user: str = Depends(get_current_user)
):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    
    try:
        # Validate
        if file.content_type not in ["image/png", "image/jpeg"]:
             raise HTTPException(status_code=400, detail="Format signature harus PNG/JPG (Transparan direkomendasikan)")

        # Process upload
        # We assume the file is already a processed image (cropped) or raw signature
        # Just save it.
        result = await process_image_upload(file, "signatures", db)
        
        signature_url = f"/api/uploads/{result['optimized']}"
        
        await db.pegawai.update_one(
            {"_id": ObjectId(id)},
            {"$set": {"signature_url": signature_url, "updated_at": datetime.now(timezone.utc)}}
        )
        
        return {"message": "Tanda tangan berhasil disimpan", "url": signature_url}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

        
    except Exception as e:
        # Catch file reading errors
        if isinstance(e, HTTPException): raise e
        print(e)
        raise HTTPException(status_code=400, detail=f"Error membaca file: {str(e)}")
    return res


@router.post("/{id}/upload-dokumen")
async def upload_pegawai_dokumen(
    id: str,
    file: UploadFile = File(...),
    keterangan: str = Form(...),
    current_user: str = Depends(get_current_user)
):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    
    # Check File Size (Approximate via content-length header if available, or read)
    # UploadFile.file is a SpooledTemporaryFile
    file.file.seek(0, 2)
    file_size = file.file.tell()
    file.file.seek(0)
    
    if file_size > 1 * 1024 * 1024: # 1MB Limit
        raise HTTPException(status_code=400, detail="Ukuran file maksimal 1MB")
        
    allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Format file harus PDF atau Gambar (JPG/PNG)")

    try:
        # Save File
        upload_dir = "/app/uploads/pegawai_docs"
        os.makedirs(upload_dir, exist_ok=True)
        
        filename = f"{uuid.uuid4()}{os.path.splitext(file.filename)[1]}"
        file_path = os.path.join(upload_dir, filename)
        
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
            
        file_url = f"/api/uploads/pegawai_docs/{filename}"
        
        # Create Document Object
        from models import PegawaiDocument
        new_doc = PegawaiDocument(
            filename=filename,
            original_name=file.filename,
            file_url=file_url,
            keterangan=keterangan,
            file_type="pdf" if "pdf" in file.content_type else "image"
        )
        
        # Update DB
        await db.pegawai.update_one(
            {"_id": ObjectId(id)},
            {"$push": {"dokumen": new_doc.dict()}}
        )
        
        return {"message": "Dokumen berhasil diupload", "data": new_doc.dict()}
        
    except Exception as e:
        print(f"Upload Error: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengupload dokumen")

@router.delete("/{id}/dokumen/{doc_id}")
async def delete_pegawai_dokumen(id: str, doc_id: str, current_user: str = Depends(get_current_user)):
    if not ObjectId.is_valid(id): raise HTTPException(status_code=400)
    
    # Pull from array
    res = await db.pegawai.update_one(
        {"_id": ObjectId(id)},
        {"$pull": {"dokumen": {"id": doc_id}}}
    )
    
    if res.modified_count == 0:
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan")
        
    # Optional: Delete physical file (Needs logic to find filename from DB before pull, or accept orphan files)
    # For now, we just remove reference to keep it simple and safe.
    
    return {"message": "Dokumen dihapus"}

# ========== KONTRAK & NOTIFIKASI ==========

@router.get("/notifications/expiring")
async def get_expiring_notifications(current_user: dict = Depends(get_current_user)):
    """Get notifications for expiring contracts and assignments"""
    from datetime import timedelta
    
    today = datetime.now(timezone.utc).date()
    one_month_later = today + timedelta(days=30)
    two_weeks_later = today + timedelta(days=14)
    one_week_later = today + timedelta(days=7)
    
    notifications = []
    
    # Get all active employees
    employees = await db.pegawai.find({"status": "AKTIF"}).to_list(None)
    
    for emp in employees:
        emp_id = str(emp["_id"])
        emp_name = emp.get("nama_lengkap", "Unknown")
        
        # Check Non-ASN contract expiry
        tgl_selesai = emp.get("tgl_selesai_kontrak")
        if tgl_selesai:
            try:
                if isinstance(tgl_selesai, str):
                    end_date = datetime.fromisoformat(tgl_selesai.replace('Z', '+00:00')).date()
                else:
                    end_date = tgl_selesai.date() if hasattr(tgl_selesai, 'date') else tgl_selesai
                
                days_remaining = (end_date - today).days
                
                if days_remaining <= 30 and days_remaining >= 0:
                    urgency = "critical" if days_remaining <= 7 else "high" if days_remaining <= 14 else "medium"
                    
                    # Check if employee has assets
                    assets = await db.aset.find({"pemegang_id": emp_id}).to_list(None)
                    has_assets = len(assets) > 0
                    
                    notifications.append({
                        "type": "contract_expiring",
                        "urgency": urgency,
                        "pegawai_id": emp_id,
                        "pegawai_nama": emp_name,
                        "nip": emp.get("nip") or emp.get("nik") or emp.get("nrp"),
                        "status_kepegawaian": emp.get("status_kepegawaian"),
                        "end_date": str(end_date),
                        "days_remaining": days_remaining,
                        "has_assets": has_assets,
                        "asset_count": len(assets),
                        "message": f"Kontrak {emp_name} akan berakhir dalam {days_remaining} hari ({end_date})"
                    })
            except Exception as e:
                print(f"Error parsing date for {emp_name}: {e}")
        
        # Check ASN assignment (penugasan) expiry
        if emp.get("status_penempatan") == "Penugasan":
            masa_penugasan_end = emp.get("masa_penugasan_end")
            if masa_penugasan_end:
                try:
                    if isinstance(masa_penugasan_end, str):
                        end_date = datetime.fromisoformat(masa_penugasan_end.replace('Z', '+00:00')).date()
                    else:
                        end_date = masa_penugasan_end.date() if hasattr(masa_penugasan_end, 'date') else masa_penugasan_end
                    
                    days_remaining = (end_date - today).days
                    
                    if days_remaining <= 30 and days_remaining >= 0:
                        urgency = "critical" if days_remaining <= 7 else "high" if days_remaining <= 14 else "medium"
                        
                        assets = await db.aset.find({"pemegang_id": emp_id}).to_list(None)
                        has_assets = len(assets) > 0
                        
                        notifications.append({
                            "type": "assignment_expiring",
                            "urgency": urgency,
                            "pegawai_id": emp_id,
                            "pegawai_nama": emp_name,
                            "nip": emp.get("nip"),
                            "status_kepegawaian": emp.get("status_kepegawaian"),
                            "end_date": str(end_date),
                            "days_remaining": days_remaining,
                            "has_assets": has_assets,
                            "asset_count": len(assets),
                            "message": f"Masa penugasan {emp_name} akan berakhir dalam {days_remaining} hari ({end_date})"
                        })
                except Exception as e:
                    print(f"Error parsing assignment date for {emp_name}: {e}")
    
    # Sort by urgency and days remaining
    urgency_order = {"critical": 0, "high": 1, "medium": 2}
    notifications.sort(key=lambda x: (urgency_order.get(x["urgency"], 3), x["days_remaining"]))
    
    return {
        "total": len(notifications),
        "critical": len([n for n in notifications if n["urgency"] == "critical"]),
        "high": len([n for n in notifications if n["urgency"] == "high"]),
        "medium": len([n for n in notifications if n["urgency"] == "medium"]),
        "notifications": notifications
    }

@router.post("/{id}/renew-contract")
async def renew_contract(
    id: str, 
    nomor_kontrak: str = Form(...),
    tgl_mulai: str = Form(...),
    tgl_selesai: str = Form(...),
    keterangan: str = Form(""),
    current_user: dict = Depends(get_current_user)
):
    """Renew/extend employee contract and save history"""
    if not ObjectId.is_valid(id): 
        raise HTTPException(status_code=400, detail="Invalid ID")
    
    pegawai = await db.pegawai.find_one({"_id": ObjectId(id)})
    if not pegawai:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    
    # Save old contract to history
    old_contract = {
        "nomor_kontrak": pegawai.get("nomor_kontrak"),
        "tgl_mulai": pegawai.get("tgl_mulai_kontrak"),
        "tgl_selesai": pegawai.get("tgl_selesai_kontrak"),
        "keterangan": "Kontrak sebelumnya",
        "renewed_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Get existing history or create new
    riwayat = pegawai.get("riwayat_kontrak", [])
    if old_contract["nomor_kontrak"] or old_contract["tgl_mulai"]:
        riwayat.append(old_contract)
    
    # Update with new contract
    await db.pegawai.update_one(
        {"_id": ObjectId(id)},
        {"$set": {
            "nomor_kontrak": nomor_kontrak,
            "tgl_mulai_kontrak": tgl_mulai,
            "tgl_selesai_kontrak": tgl_selesai,
            "riwayat_kontrak": riwayat
        }}
    )
    
    # Log activity
    await log_activity(
        db=db,
        user_id=str(current_user.id),
        user_name=current_user.full_name or "Unknown",
        action="RENEW_CONTRACT",
        module="Pegawai",
        target_id=id,
        details=f"Perpanjangan kontrak {pegawai.get('nama_lengkap')}: {nomor_kontrak}",
        metadata={"nomor_kontrak": nomor_kontrak, "tgl_mulai": tgl_mulai, "tgl_selesai": tgl_selesai}
    )
    
    return {"message": "Kontrak berhasil diperpanjang", "nomor_kontrak": nomor_kontrak}

@router.get("/{id}/contract-history")
async def get_contract_history(id: str, current_user: dict = Depends(get_current_user)):
    """Get contract renewal history for an employee"""
    if not ObjectId.is_valid(id): 
        raise HTTPException(status_code=400)
    
    pegawai = await db.pegawai.find_one({"_id": ObjectId(id)})
    if not pegawai:
        raise HTTPException(status_code=404)
    
    # Current contract
    current = {
        "nomor_kontrak": pegawai.get("nomor_kontrak"),
        "tgl_mulai": pegawai.get("tgl_mulai_kontrak"),
        "tgl_selesai": pegawai.get("tgl_selesai_kontrak"),
        "is_current": True
    }
    
    # Historical contracts
    history = pegawai.get("riwayat_kontrak", [])
    
    return {
        "pegawai_id": id,
        "nama": pegawai.get("nama_lengkap"),
        "current_contract": current,
        "history": history,
        "total_renewals": len(history)
    }

@router.get("/{id}/assets")
async def get_pegawai_assets(id: str, current_user: dict = Depends(get_current_user)):
    """Get all assets held by an employee"""
    if not ObjectId.is_valid(id): 
        raise HTTPException(status_code=400)
    
    pegawai = await db.pegawai.find_one({"_id": ObjectId(id)})
    if not pegawai:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
    
    assets = await db.aset.find({"pemegang_id": id}).to_list(None)
    
    # Transform for response
    for asset in assets:
        asset["id"] = str(asset.pop("_id"))
    
    return {
        "pegawai_id": id,
        "nama": pegawai.get("nama_lengkap"),
        "total_assets": len(assets),
        "assets": assets
    }
